import openpyxl
import sys
# 年和月
global_year = 2019
global_month = 12

# 文件名
raw_name = "raw.xlsx"
result_name = "result.xlsx"

# 第一行有数据的行号
start_offset = 3

# 筛选后的有效行号
# 不续订以外的定金有效行号列表 对应A
order_list = []
# 续订的定金有效行号列表 对应B
continue_order_list = []
# 尾款有效行号列表 对应C
tail_list = []
# 加片选片有效行号列表 对应D
photo_select_list = []
# 回家另加有效行号列表 对应E
another_list = []
# 定金中大于500的有效行号列表 对应F销售
sales_list = []

# A定金相关
col_order_date = "R"
col_order_value = "S"

# B续订相关列名, 不用于计算营业额, 用于第三部分和第四部分
col_continue_date = "BF"
col_continue_value = "BG"

# C 尾款相关列名
col_tail_date = "AU"
col_tail_value = "AV"

# D 选片加片列名
col_photo_select_date = "BB"
col_photo_select_value = "BC"

# E 回家另加列名
col_another_date = "BJ"
col_another_value = "BK"

# 以下是渠道列
col_channel_introduction = "X"
col_channel_media = "Y"
col_channel_promotion = "Z"
col_channel_type = "AB"

# 以下是人员
col_person_shooter = "AO"
col_person_cosmetician = "AQ"
col_person_assistant = "AS"

# 工资相关
# 摄影日期
col_shoot_date = "AN"
# 服务金额
col_service_value = "AX"

#  模块三所用的按照尾款支付时间挑选的符合行号列表
shoot_list = []

# BF列 续订情况 有效日期列表
shoot_continue_order_list = []

# 模块4 全局变量
# 拍摄日期列名
shoot_date_list = []


def open_excel(filename):
    if filename is None:
        print("文件名不能为空")
        raise NameError
    try:
        # print("加载工作簿: {}".format(filename))
        workbook = openpyxl.load_workbook(filename)
        return workbook
    except Exception:
        print("加载Excel文件出现错误, 请检查路径和文件是否存在.")
        exit(1)


def open_worksheet(workbook):
    # print("正在加载工作表...")
    # print("加载的工作表是:", workbook.sheetnames[0])
    return workbook[workbook.sheetnames[0]]


# 计算数据行数
def calculateMaxRow(col):
    i = 2
    while (col[i].value is not None):
        i = i + 1
    return i - 2


# 判断时间是否符合要求
def is_date_qualified(data, year, month):
    if len(data) == 0:
        return False
    if data is None:
        return False

    data_year = data.strip().split("-")
    # print(data_year)
    if len(data_year) <= 1:
        return False

    return int(data_year[0]) == year and int(data_year[1]) == month


# 通用函数: 根据传入的有效行号列表list, 列名及金额所在的列名, 工作表 计算合计数
def calculate_total_money(list, col_name, worksheet):
    # print("开始计算{}列的合计数...".format(col_name))
    # print("符合要求的行号是 {}".format(list))
    total_money = 0
    for i in list:
        value_string = worksheet[col_name + str(i)].value

        try:
            value = float(value_string)
            total_money += value
            # print("当前处理的行号是 {}, 金额是 {}".format(i, value))
        except:
            # print("当前处理的行号是 {}, 金额字符串是{}, 无法转换".format(i, value_string))
            continue
    # print("{}列的合计数是{}".format(col_name, total_money))
    return total_money


# 筛选定金中不属于续订的内容
def generate_qualified_lists_by_continued(total_rows, offset, year, month, col_name, target_list, col_continue,
                                          worksheet):
    # print("开始生成符合要求的行号列表: ")
    for i in range(offset, total_rows + offset):
        date_string = str(worksheet[col_name + str(i)].value)
        continue_string = str(worksheet[col_continue + str(i)].value).strip()

        isq = is_date_qualified(date_string, year, month) and continue_string == "续订"
        # print("当前处理的行号是: {}, 当前时间字符串是: {}, 续订字符串是:{} , 是否满足条件: {}".format(i, date_string, continue_string, isq))

        if isq:
            target_list.append(i)


def generate_qualified_lists_by_not_continued(total_rows, offset, year, month, col_name, target_list, col_continue,
                                              worksheet):
    # print("开始生成符合要求的行号列表: ")
    for i in range(offset, total_rows + offset):
        date_string = str(worksheet[col_name + str(i)].value)
        continue_string = str(worksheet[col_continue + str(i)].value).strip()

        isq = is_date_qualified(date_string, year, month) and continue_string != "续订"
        # print("当前处理的行号是: {}, 当前时间字符串是: {}, 续订字符串是:{} , 是否满足条件: {}".format(i, date_string, continue_string, isq))

        if isq:
            target_list.append(i)


# 计算定金中满足时间的行中, 金额大于等于500的合计, 并将满足条件的行号记录在sales_list中
def calculate_greater_than_500_total(list, col_name, sale_list, worksheet):
    # print("开始计算{}列的合计数...".format(col_name))
    # print("符合要求的行号是 {}".format(list))
    total_money = 0
    for i in list:
        value_string = worksheet[col_name + str(i)].value
        try:
            value = float(value_string)
            if value >= 500:
                total_money += value
                sale_list.append(i)
                # print("当前处理的行号是 {}, 金额是 {}, 被添加到销售列表中".format(i, value))
                # else:
                #     print("当前处理的行号是 {}, 金额是 {}, 小于500不做处理".format(i, value))
        except:
            # print("当前处理的行号是 {}, 金额字符串是{}, 无法转换".format(i, value_string))
            continue
    # print("{}列的合计数是{}".format(col_name, total_money))
    # print("销售列表是{}".format(sale_list))
    return total_money


# 根据有效行号, 时间, 日期, 遍历所有行, 生成ABCDE五个符合时间的行号列表
def generate_qualified_lists(total_rows, offset, year, month, col_name, target_list, worksheet):
    # print("开始生成符合要求的行号列表: ")
    for i in range(offset, total_rows + offset):
        date_string = str(worksheet[col_name + str(i)].value)
        isq = is_date_qualified(date_string, year, month)
        # print("当前处理的行号是: {}, 当前时间字符串是: {}, 是否满足条件: {}".format(i, date_string, isq))

        if isq:
            target_list.append(i)
            # print("行号结果是: {}".format(target_list))
            # print("总数是: {}".format(len(target_list)))


# -----------------------------------------营业额 模块1--------------------------------------------------------

# 将字符串转换成key_list
def cast_string_to_list(string_value):
    if string_value is None or len(string_value) ==0:
        return []

    if string_value.strip() == "/":
        return []

    string_list = string_value.strip().split("/")
    for s in string_list:
        s = s.strip()
    return string_list


# 将某一个有效的行号的金额, 按照某一个序列的属性进行分类, 修改传入的map.
# result_map 是map, list_name是渠道某个属性, row_list 是ABCDE可用行号列表, col_value_name是与ABCDE对应的金额列名
def group_by_property(result_map, list_name, row_list, col_value_name, ws):
    for i in row_list:
        value_string = str(ws[col_value_name + str(i)].value)
        # print(value_string)
        # print(value_string)
        # 尝试转换字符
        # 如果转换成功, 向MAP中放东西
        try:
            value = float(value_string)
            key_list_string = str(ws[list_name + str(i)].value)
            key_list = cast_string_to_list(key_list_string)
            # print("{}行对应的键字符串是: {}, 键列表是 {}".format(i, key_list_string, key_list))
            if len(key_list) == 0 :
                update_map("默认", value, result_map)
            elif len(key_list)==1 and key_list[0] =="None":
                update_map("默认", value, result_map)
            else:
                for key in key_list:
                    update_map(key, value, result_map)
        except :
            # 转换不成功就继续
            continue


def update_map(key, value, result_map):
    if key in result_map.keys():
        result_map[key] = result_map[key] + value
    else:
        result_map[key] = value


# 按总营业额分类
# list_name 渠道列名
def group_by_property_total_amount(list_name, ws):
    result = {}
    group_by_property(result, list_name, order_list, col_order_value, ws)
    group_by_property(result, list_name, continue_order_list, col_order_value, ws)
    group_by_property(result, list_name, tail_list, col_tail_value, ws)
    group_by_property(result, list_name, photo_select_list, col_photo_select_value, ws)
    group_by_property(result, list_name, another_list, col_another_value, ws)

    return result


# 计算字典中值的合计
def cal_map_total(dictionary):
    total = 0
    for key in dictionary.keys():
        total += dictionary[key]
    return total


# 将字典转换成字符串输出
def map_to_string(dictionary):
    result = ""
    total = cal_map_total(dictionary)
    keys_list = sorted(dictionary.keys())
    for key in keys_list:
        result = result + str(key) + " " + str(round(dictionary[key] / total * 100, 2)) + "%"
        result = result + " "
    return result


# --------------------------------------------总营业额区分  模块2-----------------------------------------------

# 将某一个有效的行号的金额, 按照某一个序列的属性进行分类, 修改传入的map.
# result_map 是map, list_name是渠道某个属性, row_list 是ABCDE可用行号列表, col_value_name是与ABCDE对应的金额列名
def group_number_by_property(result_map, list_name, row_list, col_value_name, ws):
    for i in row_list:
        value_string = ws[col_value_name + str(i)].value
        # print(value_string)
        # 尝试转换字符
        # 如果转换成功, 增加键1
        try:
            value = float(value_string)
            key_list_string = ws[list_name + str(i)].value
            key_list = cast_string_to_list(key_list_string)
            # print("{}行对应的键字符串是: {}, 键列表是 {}".format(i, key_list_string, key_list))
            if len(key_list) == 0:
                update_map_by_number("默认", result_map)
            else:
                for key in key_list:
                    update_map_by_number(key, result_map)
        except ValueError:
            # 转换不成功就继续
            continue


# 给map增加数量
def update_map_by_number(key, result_map):
    if key in result_map.keys():
        result_map[key] = result_map[key] + 1
    else:
        result_map[key] = 1


# 转换字符串根据数量
def map_to_string_with_number(dictionary):
    result = ""
    keys_list = sorted(dictionary.keys())
    for key in keys_list:
        result = result + str(key) + " " + str(dictionary[key])
        result = result + " "
    return result

# -----------------------------------------------------拍摄情况 模块3---------------------------------------

def write_basic_content(ws):
    print(ws["b2"].value)
    ws["B2"] = str(global_year)+"年" + str(global_month)+"月营业额"
    ws["B4"] = str(ws["B4"].value).replace("X",str(global_month))
    ws["B6"] = str(ws["B6"].value).replace("X",str(global_month))
    ws["B8"] = str(ws["B8"].value).replace("X",str(global_month))
    ws["B10"] = str(ws["B10"].value).replace("X",str(global_month))
    ws["B12"] = str(ws["B12"].value).replace("X",str(global_month))
    ws["B14"] = str(ws["B14"].value).replace("X",str(global_month))

    ws["E4"] = len(order_list)
    ws["E6"] = len(continue_order_list)
    ws["E8"] = len(tail_list)
    ws["E10"] = len(photo_select_list)
    ws["E12"] = len(another_list)
    ws["E14"] = len(sales_list)





def main():
    # 打开工作表
    workbook = open_excel(raw_name)
    ws = open_worksheet(workbook)

    # 使用C列作为计算总行数的依据
    colC = ws["c"]
    # 计算出总可用行数
    total_row = calculateMaxRow(colC)

    # 生成A列表
    generate_qualified_lists_by_not_continued(total_row, start_offset, global_year, global_month, col_order_date,
                                              order_list, col_channel_media, ws)
    # 生成B列表
    generate_qualified_lists_by_continued(total_row, start_offset, global_year, global_month, col_order_date,
                                          continue_order_list, col_channel_media, ws)
    # 生成C列表
    generate_qualified_lists(total_row, start_offset, global_year, global_month, col_tail_date, tail_list, ws)
    # 生成D列表
    generate_qualified_lists(total_row, start_offset, global_year, global_month, col_photo_select_date,
                             photo_select_list, ws)
    # 生成E列表
    generate_qualified_lists(total_row, start_offset, global_year, global_month, col_another_date, another_list, ws)

    # 计算相关金额
    total_money_A = calculate_total_money(order_list, col_order_value, ws)
    total_money_B = calculate_total_money(continue_order_list, col_order_value, ws)
    total_money_C = calculate_total_money(tail_list, col_tail_value, ws)
    total_money_D = calculate_total_money(photo_select_list, col_photo_select_value, ws)
    total_money_E = calculate_total_money(another_list, col_another_value, ws)
    total_money_F = calculate_greater_than_500_total(order_list, col_order_value, sales_list, ws)

    print("当前是{}年{}月".format(global_year, global_month))
    print("结果A的金额是{}, 条数是:{}".format(total_money_A, len(order_list)))
    print("结果B的金额是{}, 条数是:{}".format(total_money_B, len(continue_order_list)))
    print("结果C的金额是{}, 条数是:{}".format(total_money_C, len(tail_list)))
    print("结果D的金额是{}, 条数是:{}".format(total_money_D, len(photo_select_list)))
    print("结果E的金额是{}, 条数是:{}".format(total_money_E, len(another_list)))
    print("结果F金额是{}, 条数是:{}".format(total_money_F, len(sales_list)))
    print("总营业额是{}".format(total_money_A + total_money_B + total_money_C + total_money_D + total_money_E))
    print("---------------------------模块1结束----------------------------")
    print()
    # 按介绍人分类
    result = group_by_property_total_amount(col_channel_introduction, ws)
    # print(result)
    # print(cal_map_total(result))
    print(map_to_string(result))

    # 按媒介渠道
    result2 = group_by_property_total_amount(col_channel_media, ws)
    # print(result2)
    # print(cal_map_total(result2))
    print(map_to_string(result2))

    # 按活动分类
    result3 = group_by_property_total_amount(col_channel_promotion, ws)
    # print(result3)
    # print(cal_map_total(result3))
    print(map_to_string(result3))

    # 按类型区分
    result4 = group_by_property_total_amount(col_channel_type, ws)
    # print(result4)
    # print(cal_map_total(result4))
    print(map_to_string(result4))
    print("---------------------------模块2结束----------------------------")
    print()

    # 摄影师统计, 按尾款时间
    generate_qualified_lists(total_row, start_offset, global_year, global_month, col_tail_date, shoot_list, ws)
    # print(shoot_list)

    # 生成摄影师统计的条数
    shooter_number_map = {}
    group_number_by_property(shooter_number_map, col_person_shooter, shoot_list, col_service_value, ws)
    # print(shooter_number_map)
    # print(cal_map_total(shooter_number_map))
    print(map_to_string_with_number(shooter_number_map))

    # 生成按摄影师统计的金额
    shooter_map = {}
    group_by_property(shooter_map, col_person_shooter, shoot_list, col_service_value, ws)
    # print(shooter_map)
    # print(cal_map_total(shooter_map))
    print(map_to_string_with_number(shooter_map))

    # 生成当月续订金额 按摄影师统计的结果
    generate_qualified_lists(total_row, start_offset, global_year, global_month, col_continue_date,
                             shoot_continue_order_list, ws)
    # print(shoot_continue_order_list)

    shoot_continue_order_map = {}
    group_number_by_property(shoot_continue_order_map, col_person_shooter, shoot_continue_order_list, col_service_value,
                             ws)
    # print(shoot_continue_order_map)
    print(map_to_string_with_number(shoot_continue_order_map))

    # 生成当月续订金额按摄影师的统计
    shoot_continue_order_number_map = {}
    group_by_property(shoot_continue_order_number_map, col_person_shooter, shoot_continue_order_list, col_service_value,
                      ws)
    # print(shoot_continue_order_number_map)
    print(map_to_string_with_number(shoot_continue_order_number_map))

    print("------------------------------------模块3结束-----------------------------------------")
    print()

    # 生成符合日期的拍摄日期列名
    generate_qualified_lists(total_row, start_offset, global_year, global_month, col_shoot_date, shoot_date_list, ws)
    print("------------------------以下是按照拍摄日期和人员分类的服务金额-------------------------------")
    print(shoot_date_list)

    #1 摄影师按照 shootdate_list 及 col_service_value的 金额统计
    shoot_value_by_shoot_date_map = {}
    group_by_property(shoot_value_by_shoot_date_map, col_person_shooter, shoot_date_list, col_service_value, ws)
    print(shoot_value_by_shoot_date_map)
    print(map_to_string_with_number(shoot_value_by_shoot_date_map))

    #2 化妆师按照 shootdate_list 及 col_service_value的 金额统计
    cosmetician_value_by_shoot_date_map = {}
    group_by_property(cosmetician_value_by_shoot_date_map, col_person_cosmetician, shoot_date_list, col_service_value,
                      ws)
    print(cosmetician_value_by_shoot_date_map)
    print(map_to_string_with_number(cosmetician_value_by_shoot_date_map))

    #3 助理 shootdate_list 及 col_service_value的 金额统计
    assistant_value_by_shoot_date_map = {}
    group_by_property(assistant_value_by_shoot_date_map, col_person_assistant, shoot_date_list, col_service_value, ws)
    print(assistant_value_by_shoot_date_map)
    print(map_to_string_with_number(assistant_value_by_shoot_date_map))

    print("------------------------以下是按照选片日期和人员分类的选片加片金额-------------------------------")
    # 确认选片日期列表
    print(photo_select_list)
    #4 摄影师按照 photo_select_list  及 col_photo_select_value 的金额统计
    shoot_value_by_select_date_map = {}
    group_by_property(shoot_value_by_select_date_map, col_person_shooter, photo_select_list, col_photo_select_value, ws)
    print(shoot_value_by_select_date_map)
    print(map_to_string_with_number(shoot_value_by_select_date_map))

    #5 化妆师按照 photo_select_list 及 col_photo_select_value 金额统计
    cosmetician_value_by_select_date_map = {}
    group_by_property(cosmetician_value_by_select_date_map, col_person_cosmetician, photo_select_list,
                      col_photo_select_value, ws)
    print(cosmetician_value_by_select_date_map)
    print(map_to_string_with_number(cosmetician_value_by_select_date_map))

    #6 助理按照 photo_select_list 及 col_photo_select_value 金额统计
    assistant_value_by_select_date_map = {}
    group_by_property(assistant_value_by_select_date_map, col_person_assistant, photo_select_list,
                      col_photo_select_value, ws)
    print(assistant_value_by_select_date_map)
    print(map_to_string_with_number(assistant_value_by_select_date_map))

    print("------------------------以下是按照选片日期和人员分类的续订金额-------------------------------")
    print(photo_select_list)
    #7 摄影师按照 photo_select_list  及 col_continue_value 的金额统计
    shoot_value_by_select_date_continue_value_map = {}
    group_by_property(shoot_value_by_select_date_continue_value_map, col_person_shooter, photo_select_list,
                      col_continue_value, ws)
    print(shoot_value_by_select_date_continue_value_map)
    print(map_to_string_with_number(shoot_value_by_select_date_continue_value_map))

    #8 化妆师按照 photo_select_list 及 col_continue_value 金额统计
    cosmetician_value_by_select_date_continue_value_map = {}
    group_by_property(cosmetician_value_by_select_date_continue_value_map, col_person_cosmetician, photo_select_list,
                      col_continue_value, ws)
    print(cosmetician_value_by_select_date_continue_value_map)
    print(map_to_string_with_number(cosmetician_value_by_select_date_continue_value_map))

    #9 助理按照 photo_select_list 及 col_continue_value 金额统计
    assistant_value_by_select_date_continue_value_map = {}
    group_by_property(assistant_value_by_select_date_continue_value_map, col_person_assistant, photo_select_list,
                      col_continue_value, ws)
    print(assistant_value_by_select_date_continue_value_map)
    print(map_to_string_with_number(assistant_value_by_select_date_continue_value_map))

    # 计算除了销售基础工资之外的部分:
    bonus = len(sales_list) * 50 + (total_money_B + total_money_D) * 0.05 + (total_money_C + total_money_E) * 0.01 + (
                                                                                                                     total_money_F - 500 * len(
                                                                                                                         sales_list)) * 0.01
    print("销售基本工资之外的部分为: {}".format(bonus))
    # 开始写入excel

    result_wb = open_excel("result.xlsx")
    result_ws = open_worksheet(result_wb)

    #函数中写入文字和全局变量相关
    write_basic_content(result_ws)

    # 计算变量在主函数中写
    result_ws["D4"] = total_money_A
    result_ws["D6"] = total_money_B
    result_ws["D8"] = total_money_C
    result_ws["D10"] = total_money_D
    result_ws["D12"] = total_money_E
    result_ws["D14"] = total_money_F
    # 第二部分写入
    result_ws["C18"] = map_to_string(result)
    result_ws["C19"] = map_to_string(result2)
    result_ws["C20"] = map_to_string(result3)
    result_ws["C21"] = map_to_string(result4)
    #第三部分写入
    result_ws["C26"] = map_to_string_with_number(shooter_number_map)
    result_ws["C27"] = map_to_string_with_number(shooter_map)
    result_ws["C28"] = map_to_string_with_number(shoot_continue_order_map) + " | "+ map_to_string_with_number(shoot_continue_order_number_map)

    #第四部分写入
    result_ws["E30"] = "摄影师"
    result_ws["F30"] = "化妆师"
    result_ws["G30"] = "助理"

    result_ws["E31"] = map_to_string_with_number(shoot_value_by_shoot_date_map)
    result_ws["F31"] = map_to_string_with_number(cosmetician_value_by_shoot_date_map)
    result_ws["G31"] = map_to_string_with_number(assistant_value_by_shoot_date_map)

    result_ws["E32"] = map_to_string_with_number(shoot_value_by_select_date_map)
    result_ws["F32"] = map_to_string_with_number(cosmetician_value_by_select_date_map)
    result_ws["G32"] = map_to_string_with_number(assistant_value_by_select_date_map)

    result_ws["D33"] = map_to_string_with_number(shoot_value_by_select_date_continue_value_map)
    result_ws["E33"] = map_to_string_with_number(cosmetician_value_by_select_date_continue_value_map)
    result_ws["F33"] = map_to_string_with_number(assistant_value_by_select_date_continue_value_map)

    result_wb.save("result.xlsx")




if __name__ == '__main__':

    # if len(sys.argv) != 4:
    #     print("参数不正确")
    #
    # global_year = sys.argv[0]
    # global_month =  sys.argv[1]
    # raw_name = sys.argv[2]
    # result_name = sys.argv[3]

    main()